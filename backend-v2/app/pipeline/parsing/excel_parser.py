"""Excel(.xlsx) 解析（openpyxl）。

每个非空 sheet → 一个 HEADING（sheet 名）+ 一个结构化 TABLE DocElement
（表格形状与 docx/pdf 同路——经 table_utils 归一，首行作表头）。
不做公式重算（data_only=True 读缓存值）；空 sheet 跳过。
"""
from __future__ import annotations

import io

from openpyxl import load_workbook

from app.pipeline.parsing.doc_element import DocElement, ParseMeta
from app.pipeline.parsing.table_utils import grid_to_logical, logical_to_struct, to_text_dump


def parse_xlsx(data: bytes, file_path: str) -> tuple[list[DocElement], ParseMeta]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    elements: list[DocElement] = []
    total_tables = 0
    try:
        for ws in wb.worksheets:
            grid = [list(row) for row in ws.iter_rows(values_only=True)]
            # 空 sheet：无行或全 None 行 → 跳过
            grid = [r for r in grid if any(c is not None and str(c).strip() for c in r)]
            if not grid:
                continue
            total_tables += 1
            n_rows, n_cols = len(grid), max(len(r) for r in grid)
            # 补齐短行（None → 空串），grid_to_struct 需矩形
            grid = [[(("" if c is None else str(c)).strip()) for c in r] + [""] * (n_cols - len(r))
                    for r in grid]
            logical = grid_to_logical(grid)
            table_data, html, desc = logical_to_struct(logical, n_rows, n_cols)
            elements.append(DocElement(type="HEADING", content=str(ws.title), heading_level=1))
            elements.append(DocElement(
                type="TABLE",
                content=f"{desc}\n{to_text_dump(logical)}",
                metadata={"table_data": table_data, "table_html": html,
                          "table_description": desc, "n_rows": n_rows, "n_cols": n_cols},
            ))
    finally:
        wb.close()
    return elements, ParseMeta(file_format="xlsx", parse_engine="openpyxl",
                               total_tables=total_tables)
